from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.db.session import get_db
from app.models.asset import Asset
from app.models.resource import Resource
from app.api import deps
from app.models.user import User
from app.services.auth_audit import safe_auth_audit
import csv
import io
from typing import List, Dict
from openpyxl import load_workbook
# xlrd 1.2.0 — 解析 .xls (BIFF) 格式的唯一可用库（xlrd 2.0+ 已移除 .xls 支持，openpyxl 不支持 .xls）。
# 风险面：仅解析用户上传的 .xls 文件，不执行宏/公式。缓解措施：接口限制 10MB 上传大小 + owner/admin/operator 鉴权。
# 详见 requirements.txt 中 xlrd 条目的安全注释。
import xlrd


router = APIRouter()

# FIX: [2026-07-16 P1] 上传文件大小限制，防止 OOM
_MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10MB


async def _read_upload_with_limit(file: UploadFile, max_bytes: int = _MAX_UPLOAD_BYTES) -> bytes:
    """读取上传文件并限制大小，超过限制抛出 413。"""
    raw = await file.read(max_bytes + 1)
    if len(raw) > max_bytes:
        raise HTTPException(status_code=413, detail=f"File too large (max {max_bytes // 1024 // 1024}MB)")
    return raw


def _audit_tid(user: User) -> str:
    return (user.tenant_id or "default").strip() or "default"


def _parse_rows_from_bytes(filename: str, data: bytes) -> List[Dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    if name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(v).strip() if v is not None else "" for v in next(rows_iter)]
        except StopIteration:
            return []
        result: List[Dict[str, str]] = []
        for row in rows_iter:
            row_dict: Dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else None
                row_dict[header] = "" if value is None else str(value)
            if any(v.strip() for v in row_dict.values()):
                result.append(row_dict)
        return result

    if name.endswith(".xls"):
        book = xlrd.open_workbook(file_contents=data)
        sheet = book.sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        result: List[Dict[str, str]] = []
        for r in range(1, sheet.nrows):
            row_dict: Dict[str, str] = {}
            empty = True
            for c, header in enumerate(headers):
                if not header:
                    continue
                value = sheet.cell_value(r, c)
                text = "" if value is None else str(value)
                if text.strip():
                    empty = False
                row_dict[header] = text
            if not empty:
                result.append(row_dict)
        return result

    raise HTTPException(status_code=400, detail="Only csv, xls, xlsx files are supported")


@router.post("/import")
async def import_channels(
    file: UploadFile = File(...),
    device_id: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin", "operator"])),
):
    """
    Import channels from CSV/XLS/XLSX.

    Expected header: gb_id,name,status,node_type,civil_code,parent_gb_id,device_gb_id
    - device_gb_id can be omitted if query parameter `device_id` is provided.
    """
    raw = await _read_upload_with_limit(file)  # FIX: [2026-07-16 P1] 限制上传大小
    if not raw:
        await safe_auth_audit(
            db,
            module="channel_import",
            action="import_channels",
            source="channel_import",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=400,
            detail="empty_file",
        )
        raise HTTPException(status_code=400, detail="File is empty")

    try:
        rows = _parse_rows_from_bytes(file.filename or "", raw)
    except HTTPException:
        await safe_auth_audit(
            db,
            module="channel_import",
            action="import_channels",
            source="channel_import",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=400,
            detail="unsupported_file_type",
            extra_summary=(file.filename or "").replace(";", ".")[:120],
        )
        raise
    if not rows:
        await safe_auth_audit(
            db,
            module="channel_import",
            action="import_channels",
            source="channel_import",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=400,
            detail="no_rows",
        )
        raise HTTPException(status_code=400, detail="No valid data rows in file")

    asset_stmt = select(Asset)
    if not current_user.is_superuser:
        asset_stmt = asset_stmt.where(Asset.tenant_id == (current_user.tenant_id or "default"))
    asset_result = await db.execute(asset_stmt)
    assets = {asset.gb_id: asset for asset in asset_result.scalars().all()}

    # FIX: [2026-07-17 P0] 批量预取已存在的 Resource 记录，避免循环内 N+1 查询
    # 原：循环内对每行执行 select(Resource).where(...)，导入 N 行触发 N 次往返
    # 现：一次性查回所有 (asset_id, gb_id) 对应的 Resource，构建内存映射
    _all_gb_ids = [(row.get("gb_id") or "").strip() for row in rows if (row.get("gb_id") or "").strip()]
    _asset_ids = list({a.id for a in assets.values()})
    _exist_map: dict[tuple[str, str], Resource] = {}
    if _all_gb_ids and _asset_ids:
        _exist_stmt = select(Resource).where(
            Resource.asset_id.in_(_asset_ids),
            Resource.gb_id.in_(_all_gb_ids),
        )
        _exist_result = await db.execute(_exist_stmt)
        for _r in _exist_result.scalars().all():
            _exist_map[(_r.asset_id, _r.gb_id)] = _r

    created = 0
    updated = 0
    for row in rows:
        gb_id = (row.get("gb_id") or "").strip()
        if not gb_id:
            continue
        name = (row.get("name") or "").strip() or gb_id
        status_raw = (row.get("status") or "").strip()
        try:
            status = 1 if int(float(status_raw)) == 1 else 0
        except ValueError:
            status = 1
        node_type_raw = (row.get("node_type") or "channel").strip().lower()
        node_type = "directory" if node_type_raw == "directory" else "channel"
        civil_code = (row.get("civil_code") or "").strip() or None
        parent_gb_id = (row.get("parent_gb_id") or "").strip() or None
        device_gb_id = (row.get("device_gb_id") or "").strip() or (device_id or "").strip()

        if not device_gb_id:
            continue
        asset = assets.get(device_gb_id)
        if not asset:
            continue

        # FIX: [2026-07-17 P0] 从预取映射中查找，替代循环内 db.execute
        resource = _exist_map.get((asset.id, gb_id))
        if resource:
            resource.name = name
            resource.status = status
            resource.node_type = node_type
            resource.civil_code = civil_code
            resource.parent_gb_id = parent_gb_id
            updated += 1
        else:
            resource = Resource(
                asset_id=asset.id,
                gb_id=gb_id,
                name=name,
                status=status,
                node_type=node_type,
                civil_code=civil_code,
                parent_gb_id=parent_gb_id,
            )
            db.add(resource)
            created += 1

    await db.commit()
    fn = (file.filename or "").replace(";", ".").strip()[:120]
    await safe_auth_audit(
        db,
        module="channel_import",
        action="import_channels",
        source="channel_import",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=200,
        detail="ok",
        extra_summary=f"created={created}; updated={updated}; row_count={len(rows)}; file={fn}",
    )
    return {"created": created, "updated": updated, "total": len(rows)}


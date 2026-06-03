from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timezone
import csv
import io
from typing import List, Dict, Any
from openpyxl import load_workbook
import xlrd

from app.api import deps
from app.core.config import settings
from app.core.push_key import generate_push_key, hash_push_key
from app.db.session import get_db
from app.models.access_source import AccessSource
from app.models.push_channel import PushChannel
from app.models.resource import Resource
from app.models.user import User
from app.utils.stream_name import normalize_stream_name
from app.services.auth_audit import safe_auth_audit
from loguru import logger

router = APIRouter()


def _audit_tid(user: User) -> str:
    return (user.tenant_id or "default").strip() or "default"


class PushChannelCreate(BaseModel):
    name: str
    stream_name: str | None = None
    enabled: bool = True
    push_key_enabled: bool = True
    gb_enabled: bool = False
    gb_id: str | None = None
    gb_name: str | None = None
    gb_parent_gb_id: str | None = None


class PushChannelUpdate(BaseModel):
    name: str | None = None
    stream_name: str | None = None
    enabled: bool | None = None
    push_key_enabled: bool | None = None
    gb_enabled: bool | None = None
    gb_id: str | None = None
    gb_name: str | None = None
    gb_parent_gb_id: str | None = None


class PushChannelGbBind(BaseModel):
    gb_id: str
    gb_name: str | None = None
    gb_parent_gb_id: str | None = None


class BatchCreateItem(BaseModel):
    name: str
    stream_name: str | None = None


class PushChannelBatchCreate(BaseModel):
    items: list[BatchCreateItem]
    enabled: bool = True
    push_key_enabled: bool = True


def _public_push_key_view(prefix: str | None) -> str:
    if not prefix:
        return ""
    return f"push_{prefix}.***"


def _parse_rows_from_bytes(filename: str, data: bytes) -> List[Dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    if name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(v).strip() if v is not None else "" for v in next(rows_iter)]
        except StopIteration:
            return []
        result: List[Dict[str, str]] = []
        for row in rows_iter:
            row_dict: Dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else None
                row_dict[header] = "" if value is None else str(value)
            if any(v.strip() for v in row_dict.values()):
                result.append(row_dict)
        return result

    if name.endswith(".xls"):
        book = xlrd.open_workbook(file_contents=data)
        sheet = book.sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        result: List[Dict[str, str]] = []
        for r in range(1, sheet.nrows):
            row_dict: Dict[str, str] = {}
            empty = True
            for c, header in enumerate(headers):
                if not header:
                    continue
                value = sheet.cell_value(r, c)
                text = "" if value is None else str(value)
                if text.strip():
                    empty = False
                row_dict[header] = text
            if not empty:
                result.append(row_dict)
        return result

    raise HTTPException(status_code=400, detail="Only csv, xls, xlsx files are supported")


def _to_bool(value: Any, default: bool | None = None) -> bool | None:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "no", "n", "off"}:
        return False
    return default


async def _unbind_gb_resource(db: AsyncSession, tenant_id: str, pc: PushChannel):
    if pc.gb_resource_id:
        try:
            res = (await db.execute(select(Resource).where(Resource.id == pc.gb_resource_id, Resource.tenant_id == tenant_id))).scalars().first()
            if res and not getattr(res, "asset_id", None):
                await db.delete(res)
        except Exception as e:
            logger.warning(f"Error: {e}")
    pc.gb_enabled = False
    pc.gb_resource_id = None


async def _bind_gb_resource(
    db: AsyncSession,
    tenant_id: str,
    pc: PushChannel,
    gb_id: str,
    gb_name: str,
    gb_parent_gb_id: str | None,
):
    gb_id = (gb_id or "").strip()
    gb_name = (gb_name or "").strip() or gb_id
    gb_parent = (gb_parent_gb_id or "").strip() or None
    if pc.gb_resource_id:
        res = (await db.execute(select(Resource).where(Resource.id == pc.gb_resource_id, Resource.tenant_id == tenant_id))).scalars().first()
        if res:
            if not gb_id:
                gb_id = str(res.gb_id or "").strip()
                if not gb_id:
                    raise HTTPException(status_code=400, detail="gb_id cannot be empty when GB integration is enabled")
            if str(res.gb_id or "") != gb_id:
                exists_stmt = select(Resource).where(Resource.tenant_id == tenant_id, Resource.gb_id == gb_id, Resource.id != res.id)
                if (await db.execute(exists_stmt)).scalars().first():
                    raise HTTPException(status_code=400, detail="gb_id already exists")
            res.gb_id = gb_id
            res.name = gb_name
            res.parent_gb_id = gb_parent
            pc.gb_enabled = True
            return
    if not gb_id:
        raise HTTPException(status_code=400, detail="gb_id cannot be empty when GB integration is enabled")
    exists_stmt = select(Resource).where(Resource.tenant_id == tenant_id, Resource.gb_id == gb_id)
    if (await db.execute(exists_stmt)).scalars().first():
        raise HTTPException(status_code=400, detail="gb_id already exists")
    res = Resource(
        tenant_id=tenant_id,
        asset_id=None,
        gb_id=gb_id,
        name=gb_name,
        node_type="channel",
        parent_gb_id=gb_parent,
        status=1,
    )
    db.add(res)
    await db.flush()
    pc.gb_resource_id = res.id
    pc.gb_enabled = True


async def _get_or_init_push_channel(db: AsyncSession, tenant_id: str, source: AccessSource) -> PushChannel:
    pc = (await db.execute(select(PushChannel).where(PushChannel.id == source.id))).scalars().first()
    stream = normalize_stream_name(source.stream_name or source.name or source.id, fallback=source.id)
    if pc:
        if not pc.stream_name:
            pc.stream_name = stream
            await db.commit()
        return pc
    pc = PushChannel(
        id=source.id,
        tenant_id=tenant_id,
        stream_name=stream,
        push_key_enabled=False,
        gb_enabled=False,
        gb_resource_id=None,
    )
    db.add(pc)
    await db.commit()
    return pc


@router.get("")
async def list_push_channels(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    tenant_id = current_user.tenant_id or "default"
    sources = (
        await db.execute(
            select(AccessSource).where(
                AccessSource.tenant_id == tenant_id,
                AccessSource.protocol == "RTMP",
            )
        )
    ).scalars().all()
    out = []
    for s in sources:
        pc = await _get_or_init_push_channel(db, tenant_id, s)
        gb_id = ""
        gb_name = ""
        if pc.gb_resource_id:
            res = (
                await db.execute(select(Resource).where(Resource.id == pc.gb_resource_id, Resource.tenant_id == tenant_id))
            ).scalars().first()
            if res:
                gb_id = str(res.gb_id or "")
                gb_name = str(res.name or "")
        out.append({
            "id": s.id,
            "tenant_id": s.tenant_id,
            "name": s.name,
            "protocol": s.protocol,
            "stream_name": s.stream_name,
            "enabled": s.enabled,
            "extra": s.extra or {},
            "push_key_enabled": bool(pc.push_key_enabled),
            "push_key_hint": _public_push_key_view(pc.push_key_prefix),
            "gb_enabled": bool(pc.gb_enabled),
            "gb_resource_id": pc.gb_resource_id,
            "gb_id": gb_id,
            "gb_name": gb_name,
            "gb_stream_name": pc.stream_name,
        })
    out.sort(key=lambda x: str(x.get("name") or ""))
    return out


@router.post("")
async def create_push_channel(
    payload: PushChannelCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin"])),
):
    tenant_id = current_user.tenant_id or "default"
    name = (payload.name or "").strip()
    if not name:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="create_push_channel",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=400,
            detail="name_required",
        )
        raise HTTPException(status_code=400, detail="name cannot be empty")
    stream = normalize_stream_name(payload.stream_name or name, fallback=name)
    source = AccessSource(
        tenant_id=tenant_id,
        name=name,
        protocol="RTMP",
        host="rtmp",
        port=0,
        username=None,
        password=None,
        path="",
        stream_name=stream,
        enabled=bool(payload.enabled),
        extra={},
    )
    db.add(source)
    await db.commit()
    await db.refresh(source)

    pc = PushChannel(
        id=source.id,
        tenant_id=tenant_id,
        stream_name=stream,
        push_key_enabled=False,
        gb_enabled=False,
        gb_resource_id=None,
    )
    push_key_plain = ""
    if payload.push_key_enabled:
        push_key_plain, prefix = generate_push_key()
        pc.push_key_enabled = True
        pc.push_key_prefix = prefix
        pc.hashed_push_key = hash_push_key(push_key_plain, settings.SECRET_KEY)
    db.add(pc)

    if payload.gb_enabled:
        gb_id = (payload.gb_id or "").strip()
        gb_name = (payload.gb_name or name).strip()
        if not gb_id:
            await safe_auth_audit(
                db,
                module="push_channels",
                action="create_push_channel",
                source="push_channel_admin",
                operator=current_user.username or "unknown",
                result="failed",
                tenant_id=_audit_tid(current_user),
                status_code=400,
                detail="gb_id_required",
                extra_summary=f"channel_id={source.id}; stream_name={stream}",
            )
            raise HTTPException(status_code=400, detail="gb_id cannot be empty when GB integration is enabled")
        res = Resource(
            tenant_id=tenant_id,
            asset_id=None,
            gb_id=gb_id,
            name=gb_name,
            node_type="channel",
            parent_gb_id=(payload.gb_parent_gb_id or None),
            status=1,
        )
        db.add(res)
        await db.commit()
        await db.refresh(res)
        pc.gb_enabled = True
        pc.gb_resource_id = res.id
    await db.commit()

    await safe_auth_audit(
        db,
        module="push_channels",
        action="create_push_channel",
        source="push_channel_admin",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=201,
        detail="ok",
        extra_summary=(
            f"channel_id={source.id}; stream_name={stream}; "
            f"push_key_enabled={bool(payload.push_key_enabled)}; gb_enabled={bool(payload.gb_enabled)}"
        ),
    )
    return {
        "id": source.id,
        "stream_name": stream,
        "push_key": push_key_plain,
        "push_key_hint": _public_push_key_view(pc.push_key_prefix),
    }


@router.post("/batch")
async def batch_create_push_channels(
    payload: PushChannelBatchCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin"])),
):
    tenant_id = current_user.tenant_id or "default"
    items = payload.items or []
    if not items:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="batch_create_push_channels",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=400,
            detail="items_required",
        )
        raise HTTPException(status_code=400, detail="items cannot be empty")
    created = []
    for item in items[:200]:
        name = (item.name or "").strip()
        if not name:
            continue
        stream = normalize_stream_name(item.stream_name or name, fallback=name)
        source = AccessSource(
            tenant_id=tenant_id,
            name=name,
            protocol="RTMP",
            host="rtmp",
            port=0,
            username=None,
            password=None,
            path="",
            stream_name=stream,
            enabled=bool(payload.enabled),
            extra={},
        )
        db.add(source)
        await db.commit()
        await db.refresh(source)
        pc = PushChannel(
            id=source.id,
            tenant_id=tenant_id,
            stream_name=stream,
            push_key_enabled=False,
            gb_enabled=False,
            gb_resource_id=None,
        )
        push_key_plain = ""
        if payload.push_key_enabled:
            push_key_plain, prefix = generate_push_key()
            pc.push_key_enabled = True
            pc.push_key_prefix = prefix
            pc.hashed_push_key = hash_push_key(push_key_plain, settings.SECRET_KEY)
        db.add(pc)
        await db.commit()
        created.append({"id": source.id, "name": name, "stream_name": stream, "push_key": push_key_plain})
    await safe_auth_audit(
        db,
        module="push_channels",
        action="batch_create_push_channels",
        source="push_channel_admin",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=200,
        detail="ok",
        extra_summary=f"created_count={len(created)}; push_key_enabled={bool(payload.push_key_enabled)}",
    )
    return {"items": created}


@router.post("/import")
async def import_push_channels(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin"])),
):
    tenant_id = current_user.tenant_id or "default"
    raw = await file.read()
    if not raw:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="import_push_channels",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=400,
            detail="empty_file",
        )
        raise HTTPException(status_code=400, detail="File is empty")
    rows = _parse_rows_from_bytes(file.filename or "", raw)
    if not rows:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="import_push_channels",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=400,
            detail="no_rows",
        )
        raise HTTPException(status_code=400, detail="No valid data rows in file")

    sources = (
        await db.execute(
            select(AccessSource).where(
                AccessSource.tenant_id == tenant_id,
                AccessSource.protocol == "RTMP",
            )
        )
    ).scalars().all()
    by_stream: dict[str, AccessSource] = {}
    for s in sources:
        key = normalize_stream_name(s.stream_name or s.name or s.id, fallback=s.id)
        by_stream[key] = s

    created = 0
    updated = 0
    items: list[dict[str, str]] = []
    for row in rows[:2000]:
        name = (row.get("name") or "").strip()
        if not name:
            continue
        stream = normalize_stream_name((row.get("stream_name") or "").strip() or name, fallback=name)
        enabled = _to_bool(row.get("enabled"), default=True)
        push_key_enabled = _to_bool(row.get("push_key_enabled"), default=True)
        gb_enabled = _to_bool(row.get("gb_enabled"), default=False)
        gb_id = (row.get("gb_id") or "").strip()
        gb_name = (row.get("gb_name") or "").strip() or name
        gb_parent = (row.get("gb_parent_gb_id") or "").strip() or None

        src = by_stream.get(stream)
        push_key_plain = ""
        if not src:
            src = AccessSource(
                tenant_id=tenant_id,
                name=name,
                protocol="RTMP",
                host="rtmp",
                port=0,
                username=None,
                password=None,
                path="",
                stream_name=stream,
                enabled=bool(enabled),
                extra={},
            )
            db.add(src)
            await db.flush()
            pc = PushChannel(
                id=src.id,
                tenant_id=tenant_id,
                stream_name=stream,
                push_key_enabled=False,
                gb_enabled=False,
                gb_resource_id=None,
            )
            if push_key_enabled:
                push_key_plain, prefix = generate_push_key()
                pc.push_key_enabled = True
                pc.push_key_prefix = prefix
                pc.hashed_push_key = hash_push_key(push_key_plain, settings.SECRET_KEY)
            if gb_enabled:
                await _bind_gb_resource(db, tenant_id, pc, gb_id, gb_name, gb_parent)
            db.add(pc)
            await db.flush()
            by_stream[stream] = src
            created += 1
        else:
            src.name = name
            if enabled is not None:
                src.enabled = bool(enabled)
            await db.flush()
            pc = await _get_or_init_push_channel(db, tenant_id, src)
            if push_key_enabled is not None:
                pc.push_key_enabled = bool(push_key_enabled)
                if pc.push_key_enabled and not pc.hashed_push_key:
                    push_key_plain, prefix = generate_push_key()
                    pc.push_key_prefix = prefix
                    pc.hashed_push_key = hash_push_key(push_key_plain, settings.SECRET_KEY)
                if not pc.push_key_enabled:
                    pc.push_key_prefix = None
                    pc.hashed_push_key = None
            if gb_enabled is not None:
                if not gb_enabled:
                    await _unbind_gb_resource(db, tenant_id, pc)
                else:
                    await _bind_gb_resource(db, tenant_id, pc, gb_id or (row.get("gb_id") or ""), gb_name, gb_parent)
            pc.stream_name = normalize_stream_name(src.stream_name or src.name or src.id, fallback=src.id)
            updated += 1

        if push_key_plain:
            items.append({"name": name, "stream_name": stream, "push_key": push_key_plain})

    await db.commit()
    fn = (file.filename or "").replace(";", ".").strip()[:120]
    await safe_auth_audit(
        db,
        module="push_channels",
        action="import_push_channels",
        source="push_channel_admin",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=200,
        detail="ok",
        extra_summary=f"created={created}; updated={updated}; row_count={len(rows)}; file={fn}",
    )
    return {"created": created, "updated": updated, "total": len(rows), "items": items}


@router.post("/{channel_id}/rotate-push-key")
async def rotate_push_key(
    channel_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin"])),
):
    tenant_id = current_user.tenant_id or "default"
    src = (
        await db.execute(
            select(AccessSource).where(
                AccessSource.id == channel_id,
                AccessSource.tenant_id == tenant_id,
                AccessSource.protocol == "RTMP",
            )
        )
    ).scalars().first()
    if not src:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="rotate_push_key",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=404,
            detail="not_found",
            extra_summary=f"channel_id={channel_id}",
        )
        raise HTTPException(status_code=404, detail="Push channel not found")
    pc = await _get_or_init_push_channel(db, tenant_id, src)
    push_key_plain, prefix = generate_push_key()
    pc.push_key_enabled = True
    pc.push_key_prefix = prefix
    pc.hashed_push_key = hash_push_key(push_key_plain, settings.SECRET_KEY)
    pc.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await safe_auth_audit(
        db,
        module="push_channels",
        action="rotate_push_key",
        source="push_channel_admin",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=200,
        detail="ok",
        extra_summary=f"channel_id={channel_id}; key_prefix={prefix}",
    )
    return {"push_key": push_key_plain, "push_key_hint": _public_push_key_view(prefix)}


@router.post("/{channel_id}/save_to_gb")
async def save_to_gb(
    channel_id: str,
    payload: PushChannelGbBind,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin"])),
):
    tenant_id = current_user.tenant_id or "default"
    src = (
        await db.execute(
            select(AccessSource).where(
                AccessSource.id == channel_id,
                AccessSource.tenant_id == tenant_id,
                AccessSource.protocol == "RTMP",
            )
        )
    ).scalars().first()
    if not src:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="push_channel_bind_gb",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=404,
            detail="not_found",
            extra_summary=f"channel_id={channel_id}",
        )
        raise HTTPException(status_code=404, detail="Push channel not found")
    pc = await _get_or_init_push_channel(db, tenant_id, src)
    await _bind_gb_resource(db, tenant_id, pc, payload.gb_id, payload.gb_name or src.name, payload.gb_parent_gb_id)
    await db.commit()
    gid = (payload.gb_id or "").strip().replace(";", ".")[:80]
    await safe_auth_audit(
        db,
        module="push_channels",
        action="push_channel_bind_gb",
        source="push_channel_admin",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=200,
        detail="ok",
        extra_summary=f"channel_id={channel_id}; gb_id={gid}",
    )
    return {"ok": True}


@router.post("/{channel_id}/remove_from_gb")
async def remove_from_gb(
    channel_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin"])),
):
    tenant_id = current_user.tenant_id or "default"
    src = (
        await db.execute(
            select(AccessSource).where(
                AccessSource.id == channel_id,
                AccessSource.tenant_id == tenant_id,
                AccessSource.protocol == "RTMP",
            )
        )
    ).scalars().first()
    if not src:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="push_channel_unbind_gb",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=404,
            detail="not_found",
            extra_summary=f"channel_id={channel_id}",
        )
        raise HTTPException(status_code=404, detail="Push channel not found")
    pc = await _get_or_init_push_channel(db, tenant_id, src)
    await _unbind_gb_resource(db, tenant_id, pc)
    await db.commit()
    await safe_auth_audit(
        db,
        module="push_channels",
        action="push_channel_unbind_gb",
        source="push_channel_admin",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=200,
        detail="ok",
        extra_summary=f"channel_id={channel_id}",
    )
    return {"ok": True}


@router.put("/{channel_id}")
async def update_push_channel(
    channel_id: str,
    payload: PushChannelUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin"])),
):
    tenant_id = current_user.tenant_id or "default"
    src = (
        await db.execute(select(AccessSource).where(AccessSource.id == channel_id, AccessSource.tenant_id == tenant_id))
    ).scalars().first()
    if not src:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="update_push_channel",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=404,
            detail="not_found",
            extra_summary=f"channel_id={channel_id}",
        )
        raise HTTPException(status_code=404, detail="Push channel not found")
    if src.protocol != "RTMP":
        await safe_auth_audit(
            db,
            module="push_channels",
            action="update_push_channel",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=400,
            detail="not_rtmp",
            extra_summary=f"channel_id={channel_id}; protocol={src.protocol or ''}",
        )
        raise HTTPException(status_code=400, detail="Only RTMP supports push channel management")
    if payload.name is not None:
        src.name = payload.name.strip()
    if payload.stream_name is not None:
        src.stream_name = normalize_stream_name(payload.stream_name, fallback=src.id)
    if payload.enabled is not None:
        src.enabled = bool(payload.enabled)
    await db.commit()
    await db.refresh(src)

    pc = await _get_or_init_push_channel(db, tenant_id, src)
    if payload.push_key_enabled is not None:
        pc.push_key_enabled = bool(payload.push_key_enabled)
        if not pc.push_key_enabled:
            pc.push_key_prefix = None
            pc.hashed_push_key = None
    if payload.gb_enabled is not None and payload.gb_enabled is False:
        await _unbind_gb_resource(db, tenant_id, pc)
    else:
        need_bind = (payload.gb_enabled is True) or (payload.gb_id is not None) or (payload.gb_name is not None) or (payload.gb_parent_gb_id is not None)
        if need_bind:
            gb_id = (payload.gb_id or "").strip()
            if payload.gb_enabled is True and not gb_id and not pc.gb_resource_id:
                await safe_auth_audit(
                    db,
                    module="push_channels",
                    action="update_push_channel",
                    source="push_channel_admin",
                    operator=current_user.username or "unknown",
                    result="failed",
                    tenant_id=_audit_tid(current_user),
                    status_code=400,
                    detail="gb_id_required",
                    extra_summary=f"channel_id={channel_id}",
                )
                raise HTTPException(status_code=400, detail="gb_id cannot be empty when GB integration is enabled")
            if gb_id or pc.gb_resource_id:
                await _bind_gb_resource(db, tenant_id, pc, gb_id or (payload.gb_id or ""), (payload.gb_name or src.name), payload.gb_parent_gb_id)
    pc.stream_name = normalize_stream_name(src.stream_name or src.name or src.id, fallback=src.id)
    await db.commit()
    await safe_auth_audit(
        db,
        module="push_channels",
        action="update_push_channel",
        source="push_channel_admin",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=200,
        detail="ok",
        extra_summary=f"channel_id={channel_id}",
    )
    return {"ok": True}


@router.delete("/{channel_id}")
async def delete_push_channel(
    channel_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.require_roles(["owner", "admin"])),
):
    tenant_id = current_user.tenant_id or "default"
    src = (
        await db.execute(
            select(AccessSource).where(
                AccessSource.id == channel_id,
                AccessSource.tenant_id == tenant_id,
                AccessSource.protocol == "RTMP",
            )
        )
    ).scalars().first()
    if not src:
        await safe_auth_audit(
            db,
            module="push_channels",
            action="delete_push_channel",
            source="push_channel_admin",
            operator=current_user.username or "unknown",
            result="failed",
            tenant_id=_audit_tid(current_user),
            status_code=404,
            detail="not_found",
            extra_summary=f"channel_id={channel_id}",
        )
        raise HTTPException(status_code=404, detail="Push channel not found")
    pc = (await db.execute(select(PushChannel).where(PushChannel.id == channel_id, PushChannel.tenant_id == tenant_id))).scalars().first()
    if pc and pc.gb_resource_id:
        res = (await db.execute(select(Resource).where(Resource.id == pc.gb_resource_id, Resource.tenant_id == tenant_id))).scalars().first()
        if res and not getattr(res, "asset_id", None):
            await db.delete(res)
    if pc:
        await db.delete(pc)
    await db.delete(src)
    await db.commit()
    await safe_auth_audit(
        db,
        module="push_channels",
        action="delete_push_channel",
        source="push_channel_admin",
        operator=current_user.username or "unknown",
        result="success",
        tenant_id=_audit_tid(current_user),
        status_code=200,
        detail="ok",
        extra_summary=f"channel_id={channel_id}",
    )
    return {"ok": True}